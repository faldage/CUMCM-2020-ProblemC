import xlrd
import xlwt
import openpyxl
from openpyxl import Workbook
from enum import Enum
from openpyxl import load_workbook
import datetime

# 初始化
from xlrd import xldate_as_tuple

info = dict()
sl = dict()
# 添加
info[0] = {'农', '林', '牧', '渔', '菜', '木', '花', '大闸蟹', '猕猴桃'}
sl[0] = {11, 0}
info[1] = {'矿'}
sl[1] = {17, 0}
info[2] = {'食品', '服装', '药', '制造', '鞋', '灯', '饰', '门窗', '调味品', '家居', '化工',
           '家电', '电器', '电子', '金属', '材料', '设备', '实业', '汽车美容', '物资', '空调',
           '工艺品', '机械', '办公', '五金', '地毯', '厨房', '纺织', '卫浴', '机械', '轮胎',
           '机电', '器械', '钢', '塑料', '纸', '合金', '童装', '汽贸', '塑胶'}
sl[2] = {17, 0}
info[3] = {'电力', '热力', '燃气', '水', '天然气', '石化', '电气'}
sl[3] = {11, 0}
info[4] = {'建筑', '土木', '建设', '建材', '景观', '园艺', '路', '桥'}
sl[4] = {11, 0}
info[5] = {'批发', '零售'}
sl[5] = {17, 0}
info[6] = {'交通', '运输', '邮政', '物流', '快递', '运业', '货运', '运贸'}
sl[6] = {11, 6, 0}
info[7] = {'住宿', '餐饮'}
sl[7] = {17, 6, 0}
info[8] = {'信息', '软件', '科技', '通讯', '通信', '网络'}
sl[8] = {17, 11, 6, 0}
info[9] = {'金融', '保险', '商贸', '财务', '贸易', '发展', '个体经营', '工贸'}
sl[9] = {17, 6, 0}
info[10] = {'房地产'}
sl[10] = {11, 0}
info[11] = {'租赁', '服务', '劳务', '人力资源', '咨询', '税务', '律师', '策划', '工程检测', '质量检验测试', '招投标代理'}
sl[11] = {17, 11, 0}
info[12] = {'研究', '技术', '科技'}
sl[12] = {6, 0}
info[13] = {'水利', '生态', '环境', '公共设施', '土地', '地质', '环保'}
sl[13] = {17, 11, 6, 0}
info[14] = {'居民服务', '修理', '物业', '维修'}
sl[14] = {17, 6, 0}
info[15] = {'教育'}
sl[15] = {6, 0}
info[16] = {'卫生', '社会工作', '消防'}
sl[16] = {17, 6, 0}
info[17] = {'新闻', '出版', '广播', '电视', '电影', '录音', '文化', '艺术', '体育', '娱乐', '广告', '图书', '设计', '影城', '印'}
sl[17] = {17, 6, 0}
info[18] = {'机关', '机构', '组织', '管理'}
sl[18] = {6, 0}
info[19] = {'国际组织'}
sl[19] = {6, 0}


# 企业类
class Company:
    def __init__(self, _num, _mystr, _name, _type):
        self.num = _num
        self.mystr = _mystr
        self.name = _name
        self.type = _type
        self.dateNum = dict()

    def getIn(self, time, money):
        self.dateNum[time] = [0, 0]
        self.dateNum[time][0] = money

    def getOut(self, time, money):
        if not self.dateNum.__contains__(time):
            self.dateNum[time] = [0, 0]
        self.dateNum[time][1] = money


# 在info中查找type
def find1(_name):
    for i in range(0, 20):
        for item in info[i]:
            if _name.__contains__(item):
                return i
    return -1


wb1 = xlrd.open_workbook('fj2.xlsx')

sheet1 = wb1.sheet_by_index(0)
sheet2 = wb1.sheet_by_index(1)
sheet3 = wb1.sheet_by_index(2)

rows = sheet1.nrows  # 获取行数
cols = sheet1.ncols  # 获取列数

count = 0
companys = dict()

for i in range(1, rows):
    mystr = sheet1.cell(i, 0).value
    name = sheet1.cell(i, 1).value
    ans = find1(name)
    if ans == -1:
        count += 1
        print("fail")
    else:
        num = int(mystr.strip('E'))
        company = Company(num, mystr, name, ans)
        companys[num] = company

# for item in companys.values():
#   print("%d : %s : %s ''s type is %d" %(item.num, item.mystr, item.name, item.type))

rows = sheet2.nrows  # 获取行数
cols = sheet2.ncols  # 获取列数

i = 1
while i < rows:
    mystr = sheet2.cell(i, 0).value
    num = int(mystr.strip('E'))
    while i < rows and sheet2.cell(i, 0).value == mystr:
        temp = 0.0
        dt = sheet2.cell(i, 2).value
        dtformat = xlrd.xldate.xldate_as_datetime(dt, 0)
        while i < rows and sheet2.cell(i, 2).value == dt:
            if str(sheet2.cell(i, 7).value).__contains__('作废'):
                i += 1
            else:
                temp += float(sheet2.cell(i, 6).value)
                i += 1
        companys[num].getIn(str(dtformat.date()), temp)

rows = sheet3.nrows  # 获取行数
cols = sheet3.ncols  # 获取列数

i = 1
while i < rows:
    mystr = sheet3.cell(i, 0).value
    num = int(mystr.strip('E'))
    while i < rows and sheet3.cell(i, 0).value == mystr:
        temp = 0.0
        dt = sheet3.cell(i, 2).value
        dtformat = xlrd.xldate.xldate_as_datetime(dt, 0)
        while i < rows and sheet3.cell(i, 2).value == dt:
            if str(sheet3.cell(i, 7).value).__contains__('作废'):
                i += 1
            else:
                temp += float(sheet3.cell(i, 6).value)
                i += 1
        companys[num].getOut(str(dtformat.date()), temp)

# print("%s %s %s %s" %(companys[1].num, companys[1].mystr, companys[1].name, companys[1].type))
# for key in sorted(companys[1].dateNum.keys()):
#    print("%s : %f %f " %(key, companys[1].dateNum[key][0], companys[1].dateNum[key][1]))

# for i in range(1, 124):
#    print("----------%d" %i)
#    for key in sorted(companys[i].dateNum).keys():
#        print("%s : %f %f " %(key, companys[i].dateNum[key][0], companys[i].dateNum[key][1]))


new = Workbook()
ans = new.active
ans.cell(1, 1).value = '代号'
begin = datetime.date(2017, 1, 1)
end = datetime.date(2020, 12, 31)

no = 2
for i in range((end - begin).days + 1):
    day = begin + datetime.timedelta(days=i)
    ans.cell(1, no).value = day
    no += 1
ans.cell(1, no).value = '负数天数'
ans.cell(1, no + 1).value = 'min'

count = 0
for num in range(124, 426):
    ans.cell(num - 122, 1).value = companys[num].num
    sum = 0.0
    no = 2
    count = 0
    min = 0.0
    for i in range((end - begin).days + 1):
        day = begin + datetime.timedelta(days=i)
        if companys[num].dateNum.keys().__contains__(str(day)):
            sum -= companys[num].dateNum.get(str(day))[0]
            sum += companys[num].dateNum.get(str(day))[1]
        ans.cell(num - 122, no).value = sum
        if sum < 0:
            count += 1
        if no == 2:
            min = sum
        else:
            if min > sum:
                min = sum
        no += 1
    ans.cell(num - 122, no).value = count
    ans.cell(num - 122, no + 1).value = min

new.save(r'附件2日结.xlsx')
