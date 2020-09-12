import xlrd
import xlwt
import openpyxl
from openpyxl import Workbook
from enum import Enum
from openpyxl import load_workbook

#用于生成每个企业风险，不合格率

info = dict()
sl = dict()
ssl = {17, 11, 6, 0}
# 添加
info[0] = {'农', '林', '牧', '渔', '菜', '木', '花', '大闸蟹', '猕猴桃'}
sl[0] = {11, 0}
info[1] = {'矿'}
sl[1] = {17, 0}
info[2] = {'食品', '服装', '药', '制造', '鞋', '灯', '饰', '门窗', '调味品', '家居', '化工',
           '家电', '电器', '电子', '金属', '材料', '设备', '实业', '汽车美容', '物资', '空调',
           '工艺品', '机械', '办公', '五金', '地毯', '厨房', '纺织', '卫浴', '机械', '轮胎',
           '机电', '器械', '钢', '塑料', '纸', '合金', '童装', '汽贸', '塑胶'}
sl[2] = {17, 0}
info[3] = {'电力', '热力', '燃气', '水', '天然气', '石化', '电气'}
sl[3] = {11, 0}
info[4] = {'建筑', '土木', '建设', '建材', '景观', '园艺', '路', '桥'}
sl[4] = {11, 0}
info[5] = {'批发', '零售'}
sl[5] = {17, 0}
info[6] = {'交通', '运输', '邮政', '物流', '快递', '运业', '货运', '运贸'}
sl[6] = {11, 6, 0}
info[7] = {'住宿', '餐饮'}
sl[7] = {17, 6, 0}
info[8] = {'信息', '软件', '科技', '通讯', '通信', '网络'}
sl[8] = {17, 11, 6, 0}
info[9] = {'金融', '保险', '商贸', '财务', '贸易', '发展', '个体经营', '工贸'}
sl[9] = {17, 6, 0}
info[10] = {'房地产'}
sl[10] = {11, 0}
info[11] = {'租赁', '服务', '劳务', '人力资源', '咨询', '税务', '律师', '策划', '工程检测', '质量检验测试', '招投标代理'}
sl[11] = {17, 11, 0}
info[12] = {'研究', '技术', '科技'}
sl[12] = {6, 0}
info[13] = {'水利', '生态', '环境', '公共设施', '土地', '地质', '环保'}
sl[13] = {17, 11, 6, 0}
info[14] = {'居民服务', '修理', '物业', '维修'}
sl[14] = {17, 6, 0}
info[15] = {'教育'}
sl[15] = {6, 0}
info[16] = {'卫生', '社会工作', '消防'}
sl[16] = {17, 6, 0}
info[17] = {'新闻', '出版', '广播', '电视', '电影', '录音', '文化', '艺术', '体育', '娱乐', '广告', '图书', '设计', '影城', '印'}
sl[17] = {17, 6, 0}
info[18] = {'机关', '机构', '组织', '管理'}
sl[18] = {6, 0}
info[19] = {'国际组织'}
sl[19] = {6, 0}

class Company:
    def __init__(self, _num, _mystr,  _name, _type):
        self.num = _num
        self.mystr = _mystr
        self.name = _name
        self.type = _type
        self.dateNum = dict()

    def getIn(self, time, money):
        self.dateNum[time] = [0, 0]
        self.dateNum[time][0] = money

    def getOut(self, time, money):
        if not self.dateNum.__contains__(time):
            self.dateNum[time] = [0, 0]
        self.dateNum[time][1] = money


def find1(_name):
    for i in range(0, 20):
        for item in info[i]:
            if _name.__contains__(item):
                return i
    return -1


wb1 = xlrd.open_workbook('fj1.xlsx')

sheet1 = wb1.sheet_by_index(0)
sheet2 = wb1.sheet_by_index(1)
sheet3 = wb1.sheet_by_index(2)

rows = sheet1.nrows  # 获取行数
cols = sheet1.ncols  # 获取列数

count = 0
companys = dict()

for i in range(1, rows):
    mystr = sheet1.cell(i, 0).value
    name = sheet1.cell(i, 1).value
    ans = find1(name)
    if ans == -1:
        count += 1
        print("fail")
    else:
        num = int(mystr.strip('E'))
        company = Company(num, mystr, name, ans)
        companys[num] = company


f1_in = open('附件1进项.txt', 'r', encoding='UTF-8')
f1_out = open('附件1销项.txt', 'r', encoding='UTF-8')
#f2_in = open('附件2进项.txt', 'r', encoding='UTF-8')
#f2_out = open('附件2销项.txt', 'r', encoding='UTF-8')

f_hyfl = open('hyfl.txt', 'r', encoding ='UTF-8')

fx = dict()
while True:
    lines = f_hyfl.readline()
    if not lines:
        break
        pass
    res = lines.split(' ')
    ch = res[0]
    _type = int(ord(ch) - ord('A'))
    fx[_type] = res[2]

new = Workbook()
ans = new.active

ans.cell(1, 1).value = '代号'
ans.cell(1, 2).value = '风险'
ans.cell(1, 3).value = '进项不合格率'
ans.cell(1, 4).value = '销项不合格率'

while True:
    lines = f1_in.readline() # 整行读取数据
    if not lines:
        break
        pass
    res = lines.split(' ')
    num = int(res[0].strip('E'))
    _type = companys[num].type
    n = int(res[1])
    loc = 2
    sum = 0.0
    sumbad = 0.0
    for i in range(0, n):
        getsl = float(res[loc])
        getcount = float(res[loc + 1])
        sum += getcount
        if not ssl.__contains__(getsl):
            sumbad += getcount
        loc += 2
    ans.cell(num + 1, 1).value = res[0]
    ans.cell(num + 1, 2).value = float(fx[_type])
    ans.cell(num + 1, 3).value = (float(sumbad))/(float(sum))


while True:
    lines = f1_out.readline() # 整行读取数据
    if not lines:
        break
        pass
    res = lines.split(' ')
    num = int(res[0].strip('E'))
    _type = companys[num].type
    n = int(res[1])
    loc = 2
    sum = 0.0
    sumbad = 0.0
    for i in range(0, n):
        getsl = float(res[loc])
        getcount = float(res[loc + 1])
        sum += getcount
        if not ssl.__contains__(getsl):
            sumbad += getcount
        loc += 2
    ans.cell(num + 1, 4).value = (float(sumbad))/(float(sum))


new.save(r'附件1进项销项不合格率以及行业风险.xlsx')



