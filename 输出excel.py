import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

new = Workbook()
ans = new.active
ans.cell(1, 1).value = '企业名称'
ans.cell(1, 2).value = '进项总额'
ans.cell(1, 3).value = '销项总额'
ans.cell(1, 4).value = '销项-进项差值'

# ans.cell(1, 5).value = '进项单位平均'
# ans.cell(1, 6).value = '销项单位平均'

ans.cell(1, 5).value = '进项特殊发票数量比'
ans.cell(1, 6).value = '进项特殊发票金额比'
ans.cell(1, 7).value = '销项特殊发票数量比'
ans.cell(1, 8).value = '销项特殊发票金额比'

ans.cell(1, 9).value = '是否违约'

# --------------------进销项单位平均---------------------
with open('进销项单位平均.txt', 'r', encoding='UTF-8') as f:
    average = f.readlines()
m = 2  # 行数开始
for lines in average:
    item = lines.split(' ')
    ans.cell(m, 1).value = item[0]
    # ans.cell(m, 5).value = float(item[1])
    # ans.cell(m, 6).value = float(item[2])
    m += 1

# --------------------进销项总数------------------
with open('output1.txt', 'r', encoding='UTF-8') as f:
    inList = f.readlines()
    # 以行为分隔的数组
with open('output2.txt', 'r', encoding='UTF-8') as f:
    outList = f.readlines()
m = 2
for i in range(0, len(inList)):
    # print(inList[i].split(' ')[1])
    # print(outList[i].split(' ')[1])
    ans.cell(m, 2).value = float(inList[i].split(' ')[2])  # 进项总额
    ans.cell(m, 3).value = float(outList[i].split(' ')[2])  # 销项总额
    ans.cell(m, 4).value = float(ans.cell(m, 3).value) - \
        float(ans.cell(m, 2).value)
    m += 1

# -------------------特殊发票计算-------------------
with open('特殊发票计算.txt', 'r', encoding='UTF-8') as f:
    special = f.readlines()
m = 2
for lines in special:
    item = lines.split(' ')
    ans.cell(m, 5).value = float(item[1])
    ans.cell(m, 6).value = float(item[2])
    ans.cell(m, 7).value = float(item[3])
    ans.cell(m, 8).value = float(item[4])
    m += 1

# 是否违约
with open('是否违约.txt', 'r', encoding='UTF-8') as f:
    weiyue = f.readlines()

m = 2
for lines in weiyue:
    item = lines.split(' ')
    ans.cell(m, 9).value = int(item[1])
    m += 1

new.save(r'输出数据excel.xlsx')
